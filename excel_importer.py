"""
excel_importer.py — Importa i dati dai fogli Cassa_XXXX di un file Excel in CSV.
"""

import openpyxl
import csv
import os
from pathlib import Path
from datetime import datetime

def import_excel_data(excel_path: str, output_dir: str):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    sheets = [s for s in wb.sheetnames if s.startswith("Cassa_") and s != "Cassa_Modello"]
    
    print(f"Fogli cassa trovati: {sheets}")
    
    # ── Estrazione Movimenti (Cassa_XXXX) ──
    for sheet_name in sheets:
        year = sheet_name.split("_")[1]
        csv_name = f"cassa_{year}.csv"
        csv_path = Path(output_dir) / csv_name
        
        # Se il file esiste già, lo saltiamo per sicurezza
        if csv_path.exists():
            print(f"Salto {csv_name} (già esistente)")
            continue
            
        sheet = wb[sheet_name]
        movimenti = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            d, desc, cod, voce, imp, tipo = row[0], row[1], row[2], row[3], row[4], row[5]
            if d is None or cod is None or cod == "Z.Z.2" or cod == "Z.Z.6":
                continue
            t = "cassa" if "cassa" in str(tipo).lower() else "cc"
            if isinstance(d, datetime):
                d_str = d.strftime("%Y-%m-%d")
            else:
                d_str = str(d)[:10]
            movimenti.append([d_str, desc or "", cod, f"{float(imp or 0):.2f}", t])
            
        if movimenti:
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f)
                w.writerow(["data", "descrizione", "codice_voce", "importo", "tipo_conto"])
                w.writerows(movimenti)
            print(f"Creato {csv_name} con {len(movimenti)} movimenti.")

    # ── Estrazione Bilanci (Bilancio_XXXX) ──
    bilanci_dir = Path(output_dir) / "bilanci_storici"
    bilanci_dir.mkdir(exist_ok=True)
    
    bil_sheets = [s for s in wb.sheetnames if s.startswith("Bilancio_")]
    print(f"Fogli bilancio trovati: {bil_sheets}")
    
    for sheet_name in bil_sheets:
        year = sheet_name.split("_")[1]
        csv_path = bilanci_dir / f"bilancio_{year}.csv"
        
        sheet = wb[sheet_name]
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                w.writerow(row)
        print(f"Estratto bilancio storico {year} in {csv_path}")

if __name__ == "__main__":
    import_excel_data('RendicontoDiCassaViale2026Ok.xlsm', '.')
